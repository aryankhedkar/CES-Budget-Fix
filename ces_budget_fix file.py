"""
CES Budgets Fix Script
======================
This script fixes the monthly budget profiles for Community Energy Scheme (CES) sites.

Problem: CES was onboarded using a generic PVGIS monthly profile instead of their own
monthly generation profile. This script re-calculates all site budgets using the correct
CES monthly profile.

The script will:
1. Read site data from the CES Excel file (Property Meter Directory.xlsx)
2. Match sites to what exists in the Metris database
3. Calculate Year 1 monthly budgets using CES's actual monthly profile (commissioning year, no degradation)
4. Apply degradation (0.4% per year) to Year 1 values for Years 2-25
5. Backup existing budgets before deletion
6. Delete old budgets and insert corrected ones
7. Generate a validation report

CES Monthly Profile (from CES's own data):
- January: 2.85%, February: 5.88%, March: 8.01%, April: 11.77%
- May: 14.67%, June: 11.52%, July: 12.30%, August: 12.10%
- September: 9.57%, October: 6.03%, November: 2.78%, December: 2.51%
"""

import psycopg2
import openpyxl
import os
import csv
import json
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dotenv import load_dotenv

load_dotenv()

# ========== CONFIGURATION ==========
DB_HOST = os.getenv('DB_HOST')
DB_PORT = int(os.getenv('DB_PORT', 5432))
DB_NAME = os.getenv('DB_NAME')
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')

ACCOUNT_NAME = "Community Energy Scheme"
NUM_YEARS_TO_GENERATE = 25
DEGRADATION_RATE = 0.004  # 0.4% per year: (1-0.004)^(n-1) where n is years since commissioning

EXCEL_FILE = "Property Meter Directory.xlsx"
BACKUP_FILE = "ces_budgets_backup.csv"
VALIDATION_REPORT_FILE = "ces_budgets_validation_report.json"
OUTPUT_SQL_FILE = "ces_budgets_fix.sql"

# Batch sizes for database operations
INITIAL_BATCH_SIZE = 100
MAX_BATCH_SIZE = 500
# ===================================

# CES Monthly Profile - percentages for each month
# This is the correct profile provided by CES (Row 4 of Self Consumption Profile sheet)
# Raw percentages from CES (sum to 99.99%, normalized below)
_CES_MONTHLY_PROFILE_RAW = {
    1: 0.0285,   # January: 2.85%
    2: 0.0588,   # February: 5.88%
    3: 0.0801,   # March: 8.01%
    4: 0.1177,   # April: 11.77%
    5: 0.1467,   # May: 14.67%
    6: 0.1152,   # June: 11.52%
    7: 0.1230,   # July: 12.30%
    8: 0.1210,   # August: 12.10%
    9: 0.0957,   # September: 9.57%
    10: 0.0603,  # October: 6.03%
    11: 0.0278,  # November: 2.78%
    12: 0.0251,  # December: 2.51%
}
# Normalize to ensure exact 100% sum (fixes rounding from 2 decimal percentages)
_PROFILE_SUM = sum(_CES_MONTHLY_PROFILE_RAW.values())
CES_MONTHLY_PROFILE = {k: v / _PROFILE_SUM for k, v in _CES_MONTHLY_PROFILE_RAW.items()}

# CES Self-Consumption Profile (for reference - not stored in site_budgets table)
CES_SELF_CONSUMPTION_PROFILE = {
    1: 0.90,   # January: 90%
    2: 0.55,   # February: 55%
    3: 0.50,   # March: 50%
    4: 0.33,   # April: 33%
    5: 0.27,   # May: 27%
    6: 0.33,   # June: 33%
    7: 0.33,   # July: 33%
    8: 0.33,   # August: 33%
    9: 0.33,   # September: 33%
    10: 0.55,  # October: 55%
    11: 0.90,  # November: 90%
    12: 0.90,  # December: 90%
}


def validate_monthly_profile():
    """Verify the monthly profile sums to 100%."""
    total = sum(CES_MONTHLY_PROFILE.values())
    if abs(total - 1.0) > 0.001:
        raise ValueError(f"Monthly profile does not sum to 100%: {total * 100:.2f}%")
    print(f"✓ Monthly profile validated: {total * 100:.2f}%")


def read_excel_data(filepath: str) -> Dict[str, dict]:
    """
    Read site data from the CES Excel file.
    Returns dict: {sto_number: {commission_date, annual_generation}}
    """
    print(f"\nReading Excel file: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Onboarding source sheet']
    
    sites_data = {}
    rows_processed = 0
    rows_skipped_no_sto = 0
    rows_skipped_not_sent = 0
    rows_skipped_no_generation = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_processed += 1
        
        sto_number = row[5]  # Property STO Number (col F, index 5)
        if not sto_number:
            rows_skipped_no_sto += 1
            continue
        
        sent_to_metris = row[3]  # sent to Metris to onboard (col D)
        if not sent_to_metris or 'yes' not in str(sent_to_metris).lower():
            rows_skipped_not_sent += 1
            continue
        
        commission_date = row[6]  # Install Commission Date (col G)
        annual_generation = row[18]  # generation (col S)
        
        if not annual_generation or annual_generation == 0:
            rows_skipped_no_generation += 1
            continue
        
        # Handle datetime objects
        if hasattr(commission_date, 'strftime'):
            commission_date = commission_date.strftime('%Y-%m-%d')
        
        # Only keep the first occurrence (avoid duplicates)
        if sto_number not in sites_data:
            sites_data[sto_number] = {
                'sto_number': str(sto_number),
                'commission_date': str(commission_date) if commission_date else None,
                'annual_generation': float(annual_generation) if annual_generation else 0.0
            }
    
    print(f"  Rows processed: {rows_processed}")
    print(f"  Rows skipped (no STO): {rows_skipped_no_sto}")
    print(f"  Rows skipped (not sent): {rows_skipped_not_sent}")
    print(f"  Rows skipped (no generation): {rows_skipped_no_generation}")
    print(f"  Unique sites extracted: {len(sites_data)}")
    
    return sites_data


def get_metris_sites_from_excel(filepath: str) -> List[str]:
    """
    Read the list of sites that are in Metris from the Property Meter Directory file.
    This is the authoritative list from the 'Sites on Metris' sheet.
    
    Returns:
        List of site names (STO numbers) that are in Metris
    """
    if not os.path.exists(filepath):
        print(f"⚠ Excel file not found: {filepath}")
        print("  Will fall back to database query only")
        return []
    
    print(f"\nReading Metris sites from: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if 'Sites on Metris' not in wb.sheetnames:
        print("⚠ 'Sites on Metris' sheet not found in Excel file")
        return []
    
    ws = wb['Sites on Metris']
    sites = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        site_name = row[0]  # First column is 'name'
        if site_name:
            sites.append(str(site_name).strip())
    
    print(f"  Found {len(sites)} sites in 'Sites on Metris' sheet")
    return sites


def get_ces_sites_from_db(conn) -> Dict[str, str]:
    """
    Get all CES sites from the database.
    Returns dict: {site_name: site_id}
    """
    with conn.cursor() as cur:
        cur.execute("""
            SELECT s.id, s.name
            FROM sites s
            JOIN accounts a ON a.id = s.organization_id
            WHERE a.name ILIKE %s
            ORDER BY s.name
        """, (f'%{ACCOUNT_NAME}%',))
        return {row[1]: row[0] for row in cur.fetchall()}


def match_sites(
    excel_data: Dict[str, dict], 
    db_sites: Dict[str, str],
    metris_site_list: Optional[List[str]] = None
) -> Tuple[List[dict], List[str], List[str], List[str]]:
    """
    Match Excel sites to database sites.
    CES sites in DB are named by their STO number.
    
    Args:
        excel_data: Dict of STO number -> site data from Excel
        db_sites: Dict of site_name -> site_id from database
        metris_site_list: Optional list of site names from Metris_DB sheet (authoritative)
    
    Returns:
        - matched_sites: list of dicts with site_id, site_name, commission_date, annual_generation
        - sites_in_excel_not_db: list of STO numbers in Excel but not in DB
        - sites_in_db_not_excel: list of site names in DB but not in Excel
        - sites_in_excel_not_metris: list of STO numbers in Excel but not in Metris_DB sheet
    """
    matched = []
    sites_in_excel_not_db = []
    sites_in_excel_not_metris = []
    
    excel_sto_numbers = set(excel_data.keys())
    db_site_names = set(db_sites.keys())
    
    # If we have the Metris_DB sheet, use it as the authoritative filter
    if metris_site_list:
        metris_site_set = set(metris_site_list)
        # Filter Excel data to only sites that are in Metris_DB sheet
        excel_sto_numbers = excel_sto_numbers & metris_site_set
        sites_in_excel_not_metris = [sto for sto in excel_data.keys() if sto not in metris_site_set]
    
    for sto_number, data in excel_data.items():
        # Skip if not in Metris (if we have the list)
        if metris_site_list and sto_number not in metris_site_list:
            continue
        
        # Try direct match (site name = STO number)
        if sto_number in db_sites:
            matched.append({
                'site_id': db_sites[sto_number],
                'site_name': sto_number,
                'commission_date': data['commission_date'],
                'annual_generation': data['annual_generation']
            })
        else:
            sites_in_excel_not_db.append(sto_number)
    
    # Find DB sites not in Excel
    sites_in_db_not_excel = [name for name in db_site_names if name not in excel_sto_numbers]
    
    return matched, sites_in_excel_not_db, sites_in_db_not_excel, sites_in_excel_not_metris


def calculate_year1_monthly_budgets(annual_generation: float) -> Dict[int, float]:
    """
    Calculate Year 1 monthly generation budgets using CES profile.
    
    Args:
        annual_generation: Total annual generation in kWh
    
    Returns:
        Dict mapping month (1-12) to generation value in kWh
    """
    return {
        month: annual_generation * percentage 
        for month, percentage in CES_MONTHLY_PROFILE.items()
    }


def calculate_all_yearly_budgets(
    year1_monthly: Dict[int, float], 
    commissioning_year: int,
    num_years: int = 25,
    degradation_rate: float = DEGRADATION_RATE
) -> List[dict]:
    """
    Calculate budgets for all years applying degradation.
    
    Year 1 (commissioning year) uses the base monthly values (no degradation).
    Years 2+ apply degradation to the Year 1 values.
    
    Args:
        year1_monthly: Dict of month -> generation for Year 1 (commissioning year budget)
        commissioning_year: The year the site was commissioned
        num_years: Number of years to generate (default 25)
        degradation_rate: Annual degradation rate (default 0.4%)
    
    Returns:
        List of dicts with {year, month, generation}
    """
    budgets = []
    
    for year_offset in range(num_years):
        actual_year = commissioning_year + year_offset
        
        if year_offset == 0:
            # Year 1 (commissioning year): Use base values, no degradation
            # Formula: (1-0.004)^(n-1) where n=1 → (1-0.004)^0 = 1.0
            degradation_factor = 1.0
        else:
            # Years 2+: Apply degradation to Year 1 values
            # Formula: (1-0.004)^(n-1) where n = years since commissioning
            # Year 2 (n=2): (1-0.004)^(2-1) = (1-0.004)^1
            # Year 3 (n=3): (1-0.004)^(3-1) = (1-0.004)^2
            # Since year_offset = n - 1, we use: (1 - rate)^year_offset
            degradation_factor = (1 - degradation_rate) ** year_offset
        
        for month, base_generation in year1_monthly.items():
            budgets.append({
                'year': actual_year,
                'month': month,
                'generation': round(base_generation * degradation_factor, 2)
            })
    
    return budgets


def backup_existing_budgets(conn, site_ids: List[str], output_file: str) -> int:
    """
    Backup existing budgets for the given sites to a CSV file.
    
    Returns:
        Number of rows backed up
    """
    print(f"\nBacking up existing budgets to: {output_file}")
    
    if not site_ids:
        print("  No sites to backup")
        return 0
    
    with conn.cursor() as cur:
        # Build query with placeholders
        placeholders = ','.join(['%s'] * len(site_ids))
        cur.execute(f"""
            SELECT site_id, year, month, generation, revenue, created_at, updated_at
            FROM site_budgets
            WHERE site_id IN ({placeholders})
            ORDER BY site_id, year, month
        """, site_ids)
        
        rows = cur.fetchall()
        
        if not rows:
            print("  No existing budgets found to backup")
            return 0
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['site_id', 'year', 'month', 'generation', 'revenue', 'created_at', 'updated_at'])
            writer.writerows(rows)
        
        print(f"  Backed up {len(rows):,} rows")
        return len(rows)


def generate_sql_statements(
    matched_sites: List[dict],
    num_years: int = 25
) -> Tuple[str, dict]:
    """
    Generate SQL statements for deleting and inserting budgets.
    
    Returns:
        - SQL string
        - Stats dict
    """
    sql_lines = []
    stats = {
        'sites_processed': 0,
        'total_delete_statements': 0,
        'total_insert_rows': 0,
        'site_details': []
    }
    
    sql_lines.append("-- CES Budgets Fix SQL")
    sql_lines.append(f"-- Generated: {datetime.now().isoformat()}")
    sql_lines.append(f"-- Account: {ACCOUNT_NAME}")
    sql_lines.append(f"-- Number of sites: {len(matched_sites)}")
    sql_lines.append(f"-- Years to generate: {num_years}")
    sql_lines.append(f"-- Degradation rate: {DEGRADATION_RATE * 100}%")
    sql_lines.append("")
    sql_lines.append("BEGIN;")
    sql_lines.append("")
    
    for site in matched_sites:
        site_id = site['site_id']
        site_name = site['site_name']
        commission_date_str = site['commission_date']
        annual_generation = site['annual_generation']
        
        if not commission_date_str:
            continue
        
        try:
            commission_date = datetime.strptime(commission_date_str.split()[0], '%Y-%m-%d')
        except (ValueError, AttributeError):
            continue
        
        commissioning_year = commission_date.year
        
        # Calculate budgets
        year1_monthly = calculate_year1_monthly_budgets(annual_generation)
        all_budgets = calculate_all_yearly_budgets(year1_monthly, commissioning_year, num_years)
        
        sql_lines.append(f"-- Site: {site_name}")
        sql_lines.append(f"-- Site ID: {site_id}")
        sql_lines.append(f"-- Commissioning: {commission_date_str}")
        sql_lines.append(f"-- Annual Generation: {annual_generation:.2f} kWh")
        sql_lines.append("")
        
        # Delete existing budgets
        sql_lines.append(f"DELETE FROM site_budgets WHERE site_id = '{site_id}';")
        stats['total_delete_statements'] += 1
        sql_lines.append("")
        
        # Insert new budgets
        sql_lines.append(f"INSERT INTO site_budgets (site_id, year, month, generation)")
        sql_lines.append("VALUES")
        
        values = []
        for budget in all_budgets:
            values.append(
                f"  ('{site_id}', {budget['year']}, {budget['month']}, {budget['generation']})"
            )
        
        sql_lines.append(',\n'.join(values) + ';')
        sql_lines.append("")
        
        stats['sites_processed'] += 1
        stats['total_insert_rows'] += len(all_budgets)
        stats['site_details'].append({
            'site_name': site_name,
            'site_id': site_id,
            'commissioning_year': commissioning_year,
            'annual_generation': annual_generation,
            'rows_to_insert': len(all_budgets)
        })
    
    sql_lines.append("COMMIT;")
    
    return '\n'.join(sql_lines), stats


def execute_in_batches(
    conn,
    matched_sites: List[dict],
    num_years: int = 25,
    initial_batch_size: int = INITIAL_BATCH_SIZE,
    max_batch_size: int = MAX_BATCH_SIZE
) -> dict:
    """
    Execute budget updates in batches for safety.
    
    Returns:
        Stats dict with results
    """
    stats = {
        'sites_processed': 0,
        'sites_failed': 0,
        'total_rows_deleted': 0,
        'total_rows_inserted': 0,
        'failed_sites': []
    }
    
    batch_size = initial_batch_size
    successful_batches = 0
    
    print(f"\nProcessing {len(matched_sites)} sites in batches...")
    print(f"Starting batch size: {batch_size}")
    
    for i in range(0, len(matched_sites), batch_size):
        batch = matched_sites[i:i + batch_size]
        batch_num = i // batch_size + 1
        
        print(f"\n  Batch {batch_num}: Processing sites {i+1}-{i+len(batch)}...")
        
        try:
            with conn.cursor() as cur:
                for site in batch:
                    site_id = site['site_id']
                    site_name = site['site_name']
                    commission_date_str = site['commission_date']
                    annual_generation = site['annual_generation']
                    
                    if not commission_date_str:
                        stats['sites_failed'] += 1
                        stats['failed_sites'].append({
                            'site_name': site_name,
                            'reason': 'No commission date'
                        })
                        continue
                    
                    try:
                        commission_date = datetime.strptime(commission_date_str.split()[0], '%Y-%m-%d')
                    except (ValueError, AttributeError) as e:
                        stats['sites_failed'] += 1
                        stats['failed_sites'].append({
                            'site_name': site_name,
                            'reason': f'Invalid commission date: {e}'
                        })
                        continue
                    
                    commissioning_year = commission_date.year
                    
                    # Calculate budgets
                    year1_monthly = calculate_year1_monthly_budgets(annual_generation)
                    all_budgets = calculate_all_yearly_budgets(year1_monthly, commissioning_year, num_years)
                    
                    # Delete existing
                    cur.execute("DELETE FROM site_budgets WHERE site_id = %s", (site_id,))
                    stats['total_rows_deleted'] += cur.rowcount
                    
                    # Insert new
                    for budget in all_budgets:
                        cur.execute(
                            """INSERT INTO site_budgets (site_id, year, month, generation)
                               VALUES (%s, %s, %s, %s)""",
                            (site_id, budget['year'], budget['month'], budget['generation'])
                        )
                        stats['total_rows_inserted'] += 1
                    
                    stats['sites_processed'] += 1
                
                conn.commit()
                successful_batches += 1
                print(f"    ✓ Batch {batch_num} committed successfully")
                
                # Scale up batch size after successful batches
                if successful_batches >= 2 and batch_size < max_batch_size:
                    batch_size = min(batch_size * 2, max_batch_size)
                    print(f"    Scaling up batch size to: {batch_size}")
        
        except Exception as e:
            conn.rollback()
            print(f"    ✗ Batch {batch_num} failed: {e}")
            print(f"    Rolling back this batch...")
            
            # Mark all sites in this batch as failed
            for site in batch:
                stats['sites_failed'] += 1
                stats['failed_sites'].append({
                    'site_name': site['site_name'],
                    'reason': f'Batch failed: {e}'
                })
            
            # Reduce batch size on failure
            if batch_size > initial_batch_size:
                batch_size = max(batch_size // 2, initial_batch_size)
                print(f"    Reducing batch size to: {batch_size}")
    
    return stats


def generate_validation_report(
    matched_sites: List[dict],
    sites_in_excel_not_db: List[str],
    sites_in_db_not_excel: List[str],
    sites_in_excel_not_metris: List[str],
    execution_stats: Optional[dict] = None
) -> dict:
    """
    Generate a comprehensive validation report.
    """
    report = {
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'total_sites_in_excel': len(matched_sites) + len(sites_in_excel_not_db) + len(sites_in_excel_not_metris),
            'total_sites_in_db': len(matched_sites) + len(sites_in_db_not_excel),
            'sites_matched': len(matched_sites),
            'sites_in_excel_not_db': len(sites_in_excel_not_db),
            'sites_in_db_not_excel': len(sites_in_db_not_excel),
            'sites_in_excel_not_metris': len(sites_in_excel_not_metris),
        },
        'configuration': {
            'degradation_rate': DEGRADATION_RATE,
            'years_generated': NUM_YEARS_TO_GENERATE,
            'monthly_profile': {str(k): v for k, v in CES_MONTHLY_PROFILE.items()},
        },
        'unmatched_sites': {
            'in_excel_not_db': sites_in_excel_not_db[:50],  # Limit to first 50
            'in_db_not_excel': sites_in_db_not_excel[:50],
            'in_excel_not_metris': sites_in_excel_not_metris[:50],
        },
        'sample_calculations': []
    }
    
    # Add sample calculations for spot-checking
    sample_sites = matched_sites[:5]
    for site in sample_sites:
        if not site['commission_date']:
            continue
        
        try:
            commission_date = datetime.strptime(site['commission_date'].split()[0], '%Y-%m-%d')
        except (ValueError, AttributeError):
            continue
        
        year1_monthly = calculate_year1_monthly_budgets(site['annual_generation'])
        year1_total = sum(year1_monthly.values())
        
        # Calculate commissioning year window (commission month to commission month - 1)
        report['sample_calculations'].append({
            'site_name': site['site_name'],
            'annual_generation': site['annual_generation'],
            'commission_date': site['commission_date'],
            'year1_total': round(year1_total, 2),
            'matches_annual': abs(year1_total - site['annual_generation']) < 0.01,
            'year1_monthly': {str(k): round(v, 2) for k, v in year1_monthly.items()},
            'year2_monthly_sample': {
                str(k): round(v * (1 - DEGRADATION_RATE), 2) 
                for k, v in list(year1_monthly.items())[:3]
            }
        })
    
    if execution_stats:
        report['execution_stats'] = execution_stats
    
    return report


def print_summary(matched_sites, sites_in_excel_not_db, sites_in_db_not_excel, sites_in_excel_not_metris=None):
    """Print a summary of matched and unmatched sites."""
    print("\n" + "=" * 60)
    print("SITE MATCHING SUMMARY")
    print("=" * 60)
    print(f"Sites matched:              {len(matched_sites):,}")
    print(f"In Excel but not in DB:     {len(sites_in_excel_not_db):,}")
    print(f"In DB but not in Excel:     {len(sites_in_db_not_excel):,}")
    if sites_in_excel_not_metris:
        print(f"In Excel but not in Metris: {len(sites_in_excel_not_metris):,}")
    
    if sites_in_excel_not_db:
        print(f"\n⚠ First 10 sites in Excel but not in DB:")
        for sto in sites_in_excel_not_db[:10]:
            print(f"    - {sto}")
        if len(sites_in_excel_not_db) > 10:
            print(f"    ... and {len(sites_in_excel_not_db) - 10} more")
    
    if sites_in_excel_not_metris:
        print(f"\n⚠ First 10 sites in Excel but not in 'Sites on Metris' sheet:")
        for sto in sites_in_excel_not_metris[:10]:
            print(f"    - {sto}")
        if len(sites_in_excel_not_metris) > 10:
            print(f"    ... and {len(sites_in_excel_not_metris) - 10} more")
    
    total_records = len(matched_sites) * 12 * NUM_YEARS_TO_GENERATE
    print(f"\nExpected operations:")
    print(f"  - DELETE budgets for {len(matched_sites):,} sites")
    print(f"  - INSERT {total_records:,} new budget records")
    print(f"  - ({len(matched_sites)} sites × 12 months × {NUM_YEARS_TO_GENERATE} years)")


def main():
    print("=" * 60)
    print("CES Budgets Fix Script")
    print("=" * 60)
    print(f"\nTarget Account: {ACCOUNT_NAME}")
    print(f"Excel File: {EXCEL_FILE}")
    print(f"Years to generate: {NUM_YEARS_TO_GENERATE}")
    print(f"Degradation rate: {DEGRADATION_RATE * 100}%")
    
    # Validate profile
    validate_monthly_profile()
    
    print("\nSelect mode:")
    print("1. Analyze only (no database changes)")
    print("2. Generate SQL file (review before executing)")
    print("3. Execute with backup (recommended)")
    print("4. List all CES sites in database")
    print("5. Generate validation report only")
    
    mode = input("\nEnter mode (1-5): ").strip()
    
    if mode == "1":
        # Analyze only - no DB connection needed for Excel parsing
        excel_data = read_excel_data(EXCEL_FILE)
        
        # Read Metris sites from Excel file
        metris_site_list = get_metris_sites_from_excel(EXCEL_FILE)
        
        print("\nConnecting to database to match sites...")
        try:
            with psycopg2.connect(
                host=DB_HOST, port=DB_PORT, database=DB_NAME,
                user=DB_USER, password=DB_PASSWORD
            ) as conn:
                db_sites = get_ces_sites_from_db(conn)
                print(f"Found {len(db_sites)} sites for account '{ACCOUNT_NAME}'")
                
                matched, not_in_db, not_in_excel, not_in_metris = match_sites(
                    excel_data, db_sites, metris_site_list
                )
                print_summary(matched, not_in_db, not_in_excel, not_in_metris)
                
        except psycopg2.Error as e:
            print(f"\nDatabase error: {e}")
        return
    
    if mode == "4":
        print("\nConnecting to database...")
        try:
            with psycopg2.connect(
                host=DB_HOST, port=DB_PORT, database=DB_NAME,
                user=DB_USER, password=DB_PASSWORD
            ) as conn:
                db_sites = get_ces_sites_from_db(conn)
                print(f"\nSites found for '{ACCOUNT_NAME}': {len(db_sites)}")
                print("-" * 50)
                for i, (name, site_id) in enumerate(sorted(db_sites.items())[:50]):
                    print(f"  {name}")
                if len(db_sites) > 50:
                    print(f"  ... and {len(db_sites) - 50} more")
        except psycopg2.Error as e:
            print(f"\nDatabase error: {e}")
        return
    
    # All other modes need both Excel data and DB connection
    excel_data = read_excel_data(EXCEL_FILE)
    
    # Read Metris sites from Excel file
    metris_site_list = get_metris_sites_from_excel(EXCEL_FILE)
    
    print("\nConnecting to database...")
    try:
        with psycopg2.connect(
            host=DB_HOST, port=DB_PORT, database=DB_NAME,
            user=DB_USER, password=DB_PASSWORD
        ) as conn:
            print("Connected successfully.")
            
            db_sites = get_ces_sites_from_db(conn)
            print(f"Found {len(db_sites)} sites for account '{ACCOUNT_NAME}'")
            
            matched, not_in_db, not_in_excel, not_in_metris = match_sites(
                excel_data, db_sites, metris_site_list
            )
            print_summary(matched, not_in_db, not_in_excel, not_in_metris)
            
            if mode == "5":
                # Generate validation report only
                report = generate_validation_report(matched, not_in_db, not_in_excel, not_in_metris)
                with open(VALIDATION_REPORT_FILE, 'w', encoding='utf-8') as f:
                    json.dump(report, f, indent=2)
                print(f"\n✓ Validation report saved to: {VALIDATION_REPORT_FILE}")
                return
            
            if mode == "2":
                # Generate SQL file
                sql, stats = generate_sql_statements(matched, NUM_YEARS_TO_GENERATE)
                with open(OUTPUT_SQL_FILE, 'w', encoding='utf-8') as f:
                    f.write(sql)
                print(f"\n✓ SQL file generated: {OUTPUT_SQL_FILE}")
                print(f"  Sites processed: {stats['sites_processed']}")
                print(f"  Delete statements: {stats['total_delete_statements']}")
                print(f"  Insert rows: {stats['total_insert_rows']:,}")
                print(f"\nReview the file, then run manually:")
                print(f"  psql -h $DB_HOST -U $DB_USER -d $DB_NAME -f {OUTPUT_SQL_FILE}")
                return
            
            if mode == "3":
                # Execute with backup
                print("\n" + "=" * 60)
                print("PRE-EXECUTION CHECKS")
                print("=" * 60)
                
                if not matched:
                    print("\n⚠ No sites matched. Nothing to do.")
                    return
                
                # Step 1: Backup
                print("\nStep 1: Creating backup of existing budgets...")
                site_ids = [s['site_id'] for s in matched]
                rows_backed_up = backup_existing_budgets(conn, site_ids, BACKUP_FILE)
                
                if rows_backed_up > 0:
                    print(f"✓ Backup saved to: {BACKUP_FILE}")
                else:
                    print("  No existing budgets found (fresh insert)")
                
                # Confirmation
                total_records = len(matched) * 12 * NUM_YEARS_TO_GENERATE
                print("\n" + "=" * 60)
                print("CONFIRMATION REQUIRED")
                print("=" * 60)
                print(f"\nThis will:")
                print(f"  1. DELETE all existing budgets for {len(matched):,} sites")
                print(f"  2. INSERT {total_records:,} new budget records")
                print(f"\nBackup file: {BACKUP_FILE}")
                
                confirm = input("\nType 'YES' to proceed: ").strip()
                if confirm != "YES":
                    print("Aborted. No changes made.")
                    return
                
                # Step 2: Execute
                print("\n" + "=" * 60)
                print("EXECUTING")
                print("=" * 60)
                
                stats = execute_in_batches(conn, matched, NUM_YEARS_TO_GENERATE)
                
                # Step 3: Generate report
                report = generate_validation_report(matched, not_in_db, not_in_excel, not_in_metris, stats)
                with open(VALIDATION_REPORT_FILE, 'w', encoding='utf-8') as f:
                    json.dump(report, f, indent=2)
                
                print("\n" + "=" * 60)
                print("EXECUTION COMPLETE")
                print("=" * 60)
                print(f"Sites processed:    {stats['sites_processed']:,}")
                print(f"Sites failed:       {stats['sites_failed']:,}")
                print(f"Rows deleted:       {stats['total_rows_deleted']:,}")
                print(f"Rows inserted:      {stats['total_rows_inserted']:,}")
                print(f"\nValidation report:  {VALIDATION_REPORT_FILE}")
                print(f"Backup file:        {BACKUP_FILE}")
                
                if stats['failed_sites']:
                    print(f"\n⚠ Failed sites:")
                    for failed in stats['failed_sites'][:10]:
                        print(f"    - {failed['site_name']}: {failed['reason']}")
                    if len(stats['failed_sites']) > 10:
                        print(f"    ... and {len(stats['failed_sites']) - 10} more (see report)")
                
                return
    
    except psycopg2.Error as e:
        print(f"\nDatabase error: {e}")
        return


if __name__ == "__main__":
    main()
